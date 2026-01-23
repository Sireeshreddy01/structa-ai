"""Spreadsheet Export Service (Excel, CSV)"""
from typing import List, Dict, Any, Optional
from io import BytesIO, StringIO
import csv
import structlog

logger = structlog.get_logger()


class SpreadsheetExporter:
    """Export structured documents to Excel and CSV formats"""
    
    def __init__(self):
        self._openpyxl = None
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            self._openpyxl = {
                'Workbook': openpyxl.Workbook,
                'Font': Font,
                'Alignment': Alignment,
                'PatternFill': PatternFill,
                'Border': Border,
                'Side': Side,
                'get_column_letter': get_column_letter,
            }
            logger.info("openpyxl loaded for Excel export")
        except ImportError:
            logger.warning("openpyxl not available, Excel export disabled")
    
    def export_excel(
        self,
        document: Dict[str, Any],
        output_path: Optional[str] = None,
        include_raw_text: bool = False,
    ) -> bytes:
        """Export structured document to Excel"""
        if not self._openpyxl:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")
        
        logger.info("Exporting document to Excel")
        
        Workbook = self._openpyxl['Workbook']
        Font = self._openpyxl['Font']
        Alignment = self._openpyxl['Alignment']
        PatternFill = self._openpyxl['PatternFill']
        Border = self._openpyxl['Border']
        Side = self._openpyxl['Side']
        get_column_letter = self._openpyxl['get_column_letter']
        
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        default_sheet.title = "Summary"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        
        # Summary sheet
        ws = default_sheet
        ws['A1'] = 'Document Summary'
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        if document.get('title'):
            ws[f'A{row}'] = 'Title'
            ws[f'B{row}'] = document['title']
            row += 1
        
        # Key-values
        if document.get('key_values'):
            row += 1
            ws[f'A{row}'] = 'Extracted Data'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for key, value in document['key_values'].items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                row += 1
        
        # Metadata
        if document.get('metadata'):
            row += 1
            ws[f'A{row}'] = 'Metadata'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            meta = document['metadata']
            for key, value in meta.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                row += 1
        
        # Auto-size columns
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 30
        
        # Tables sheets
        for i, table in enumerate(document.get('tables', [])):
            ws = wb.create_sheet(title=f"Table {i + 1}")
            
            matrix = table.get('matrix', [])
            headers = table.get('headers', [])
            
            # Write headers if available
            if headers:
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                start_row = 2
            else:
                start_row = 1
            
            # Write data
            for row_idx, row_data in enumerate(matrix, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
            
            # Auto-size columns
            for col in range(1, len(matrix[0]) + 1 if matrix else 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Raw text sheet
        if include_raw_text and document.get('raw_text'):
            ws = wb.create_sheet(title="Raw Text")
            ws['A1'] = document['raw_text']
            ws['A1'].alignment = Alignment(wrap_text=True)
            ws.column_dimensions['A'].width = 100
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        # Optionally save to file
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(excel_bytes)
            logger.info(f"Excel saved to {output_path}")
        
        logger.info("Excel export complete", size_bytes=len(excel_bytes))
        return excel_bytes
    
    def export_csv(
        self,
        document: Dict[str, Any],
        output_path: Optional[str] = None,
        table_index: int = 0,
    ) -> str:
        """Export a table from the document to CSV"""
        logger.info("Exporting to CSV", table_index=table_index)
        
        tables = document.get('tables', [])
        
        if not tables:
            # Export key-values as CSV
            if document.get('key_values'):
                buffer = StringIO()
                writer = csv.writer(buffer)
                writer.writerow(['Key', 'Value'])
                
                for key, value in document['key_values'].items():
                    writer.writerow([key.replace('_', ' ').title(), str(value)])
                
                csv_content = buffer.getvalue()
                buffer.close()
            else:
                csv_content = ""
        else:
            if table_index >= len(tables):
                table_index = 0
            
            table = tables[table_index]
            matrix = table.get('matrix', [])
            
            buffer = StringIO()
            writer = csv.writer(buffer)
            
            for row in matrix:
                writer.writerow(row)
            
            csv_content = buffer.getvalue()
            buffer.close()
        
        # Optionally save to file
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(csv_content)
            logger.info(f"CSV saved to {output_path}")
        
        logger.info("CSV export complete")
        return csv_content
    
    def export_all_tables_csv(
        self,
        document: Dict[str, Any],
        output_dir: str,
    ) -> List[str]:
        """Export all tables as separate CSV files"""
        import os
        
        tables = document.get('tables', [])
        file_paths = []
        
        for i, table in enumerate(tables):
            file_path = os.path.join(output_dir, f"table_{i + 1}.csv")
            self.export_csv(document, output_path=file_path, table_index=i)
            file_paths.append(file_path)
        
        return file_paths


# Singleton instance
spreadsheet_exporter = SpreadsheetExporter()
